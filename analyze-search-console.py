#!/usr/bin/env python3
"""Analyze Search Console Performance data - Full Diagnostic"""

import openpyxl
import sys
from pathlib import Path
from statistics import mean, stdev

excel_file = Path("seach-console-perf/lescalculateurs.fr-Performance-on-Search-2026-05-02.xlsx")

if not excel_file.exists():
    print(f"❌ File not found: {excel_file}")
    sys.exit(1)

try:
    wb = openpyxl.load_workbook(excel_file)
    print("\n" + "="*120)
    print("📊 SEARCH CONSOLE DIAGNOSTIC - lescalculateurs.fr")
    print("="*120 + "\n")
    
    # GRAPHIQUE - Données par jour
    print("📈 1️⃣ DAILY PERFORMANCE SUMMARY")
    print("-" * 120)
    ws_graph = wb['Graphique']
    data_graph = []
    for row in ws_graph.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        data_graph.append(row)
    
    if data_graph:
        clics = [r[1] for r in data_graph if r[1] is not None]
        impressions = [r[2] for r in data_graph if r[2] is not None]
        ctr = [r[3] for r in data_graph if r[3] is not None]
        position = [r[4] for r in data_graph if r[4] is not None]
        
        print(f"  📅 Period: {data_graph[0][0]} to {data_graph[-1][0]} ({len(data_graph)} days)")
        print(f"  👆 Clicks: {sum(clics):.0f} total | avg {mean(clics):.1f}/day | range {min(clics):.0f}-{max(clics):.0f}")
        print(f"  👁️ Impressions: {sum(impressions):.0f} total | avg {mean(impressions):.0f}/day")
        print(f"  🎯 CTR: avg {mean(ctr)*100:.2f}% | range {min(ctr)*100:.2f}%-{max(ctr)*100:.2f}%")
        print(f"  📍 Position: avg {mean(position):.1f} | range {min(position):.1f}-{max(position):.1f} (lower=better)")
    print()
    
    # REQUÊTES - Top queries
    print("🔍 2️⃣ TOP KEYWORDS (REQUÊTES)")
    print("-" * 120)
    ws_req = wb['Requêtes']
    req_data = []
    for row in ws_req.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        req_data.append(row)
    
    req_data.sort(key=lambda x: x[1] if x[1] else 0, reverse=True)
    print(f"  Total keywords: {len(req_data)}")
    print(f"\n  🏆 Top 15 Keywords by Clicks:")
    for i, row in enumerate(req_data[:15], 1):
        query, clics, impr, ctr_val, pos = row[0], row[1] or 0, row[2] or 0, row[3] or 0, row[4] or 0
        print(f"    {i:2d}. {query:50s} | Clics: {clics:3.0f} | Impr: {impr:5.0f} | CTR: {ctr_val*100:5.2f}% | Pos: {pos:4.1f}")
    print()
    
    # PAGES - Top pages
    print("📄 3️⃣ TOP PAGES")
    print("-" * 120)
    ws_pages = wb['Pages']
    pages_data = []
    for row in ws_pages.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        pages_data.append(row)
    
    pages_data.sort(key=lambda x: x[1] if x[1] else 0, reverse=True)
    print(f"  Total pages: {len(pages_data)}")
    print(f"\n  🏆 Top 15 Pages by Clicks:")
    for i, row in enumerate(pages_data[:15], 1):
        page, clics, impr, ctr_val, pos = row[0], row[1] or 0, row[2] or 0, row[3] or 0, row[4] or 0
        page_short = page[:65] if len(str(page)) > 65 else page
        print(f"    {i:2d}. {str(page_short):67s} | {clics:3.0f} clics | {impr:5.0f} impr | {ctr_val*100:5.2f}% CTR | {pos:4.1f}")
    print()
    
    # APPARENCE - Featured snippets, rich results
    print("✨ 4️⃣ SEARCH APPEARANCE (Rich Results, Featured Snippets, etc.)")
    print("-" * 120)
    try:
        ws_appear = wb['Apparence dans les résultats de']
        appear_data = []
        for row in ws_appear.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                break
            appear_data.append(row)
        
        if appear_data:
            appear_data.sort(key=lambda x: x[1] if x[1] else 0, reverse=True)
            for i, row in enumerate(appear_data[:10], 1):
                app_type, clics, impr, ctr_val, pos = row[0], row[1] or 0, row[2] or 0, row[3] or 0, row[4] or 0
                print(f"    {i:2d}. {str(app_type):50s} | {clics:3.0f} clics | {impr:5.0f} impr | {ctr_val*100:5.2f}% CTR")
        else:
            print("    ℹ️ No rich results data available")
    except:
        print("    ℹ️ Sheet 'Apparence' not available")
    print()
    
    # APPAREILS - Device breakdown
    print("📱 5️⃣ DEVICES")
    print("-" * 120)
    try:
        ws_dev = wb['Appareils']
        dev_data = []
        for row in ws_dev.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                break
            dev_data.append(row)
        
        dev_data.sort(key=lambda x: x[1] if x[1] else 0, reverse=True)
        total_clicks = sum([r[1] or 0 for r in dev_data])
        for row in dev_data:
            device, clics, impr, ctr_val, pos = row[0], row[1] or 0, row[2] or 0, row[3] or 0, row[4] or 0
            pct = (clics / total_clicks * 100) if total_clicks > 0 else 0
            print(f"    {device:20s} | {clics:3.0f} clics ({pct:5.1f}%) | {impr:5.0f} impr | {ctr_val*100:5.2f}% CTR | {pos:4.1f}")
    except:
        print("    ℹ️ Device data not available")
    print()
    
    print("="*120)
    print("✓ Diagnostic complete!\n")
    
except Exception as e:
    import traceback
    print(f"❌ Error: {e}")
    traceback.print_exc()
    sys.exit(1)
