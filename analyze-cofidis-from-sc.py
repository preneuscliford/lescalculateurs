#!/usr/bin/env python3
"""Analyze updated Search Console data (May 3, 2026) for COFIDIS integration planning"""

import openpyxl
from pathlib import Path
from collections import defaultdict

print("\n" + "="*180)
print("📊 COFIDIS AFFILIATION - Updated Search Console Analysis (May 3, 2026)")
print("="*180 + "\n")

excel_file = Path("seach-console-perf/lescalculateurs.fr-Performance-on-Search-2026-05-03.xlsx")

if not excel_file.exists():
    print(f"❌ File not found: {excel_file}")
    exit(1)

try:
    wb = openpyxl.load_workbook(excel_file)
    print(f"✅ Loaded workbook with sheets: {wb.sheetnames}\n")
    
    # Get Pages sheet
    if 'Pages' not in wb.sheetnames:
        print("❌ 'Pages' sheet not found. Available sheets:", wb.sheetnames)
        exit(1)
    
    ws_pages = wb['Pages']
    
    # Extract relevant pages for COFIDIS
    target_pages = {
        '/pages/pret': 'Simulateur Prêt Immobilier',
        '/pages/financement': 'Simulateur Financement',
        '/pages/charges': 'Charges Copropriété',
        '/pages/notaire': 'Frais Notaire',
        '/pages/impot': 'Simulateur Impôt',
        '/pages/salaire': 'Calculateur Salaire',
    }
    
    pages_data = {}
    
    for row_idx, row in enumerate(ws_pages.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            break
        
        page_url = row[0]
        clics = row[1] or 0
        impr = row[2] or 0
        ctr = row[3] or 0
        pos = row[4] or 0
        
        # Check if this is a target page
        for target_key, label in target_pages.items():
            if target_key in str(page_url):
                pages_data[label] = {
                    'url': page_url,
                    'clics': clics,
                    'impr': impr,
                    'ctr': ctr,
                    'pos': pos,
                }
                break
    
    # Display results
    print("="*180)
    print("🎯 COFIDIS INTEGRATION PAGES - Traffic Analysis")
    print("="*180 + "\n")
    
    for page_label, data in sorted(pages_data.items(), key=lambda x: x[1]['clics'], reverse=True):
        print(f"📄 {page_label}")
        print(f"   URL: {data['url']}")
        print(f"   Clics: {data['clics']:.0f} | Impressions: {data['impr']:.0f} | CTR: {data['ctr']*100:.2f}% | Pos: {data['pos']:.1f}")
        
        # Estimate leads with different conversion rates
        conv_low = data['clics'] * 0.02  # 2% conservative
        conv_mid = data['clics'] * 0.035  # 3.5% realistic
        conv_high = data['clics'] * 0.05  # 5% optimistic
        
        print(f"   💰 Monthly leads estimate: {conv_low:.0f}-{conv_high:.0f} (conservative: {conv_low:.0f})")
        print()
    
    # Calculate totals
    print("="*180)
    print("📈 PROJECTED MONTHLY LEADS (Current Traffic)")
    print("="*180 + "\n")
    
    if 'Simulateur Prêt Immobilier' in pages_data and 'Simulateur Financement' in pages_data:
        pret_clics = pages_data['Simulateur Prêt Immobilier']['clics']
        financement_clics = pages_data['Simulateur Financement']['clics']
        
        total_high_priority = pret_clics + financement_clics
        
        leads_conservative = total_high_priority * 0.02
        leads_realistic = total_high_priority * 0.035
        leads_optimistic = total_high_priority * 0.05
        
        print(f"🟢 HIGH PRIORITY PAGES (Prêt + Financement):")
        print(f"   Combined monthly clicks: {total_high_priority:.0f}")
        print(f"   Estimated leads range: {leads_conservative:.0f} - {leads_optimistic:.0f}/month")
        print(f"   Realistic estimate: {leads_realistic:.0f} leads/month")
        print(f"   At 45€/lead: €{leads_realistic * 45:,.0f}/month revenue potential\n")
    
    if 'Charges Copropriété' in pages_data and 'Frais Notaire' in pages_data:
        charges_clics = pages_data['Charges Copropriété']['clics']
        notaire_clics = pages_data['Frais Notaire']['clics']
        
        total_medium = charges_clics + notaire_clics
        
        leads_conservative = total_medium * 0.015
        leads_realistic = total_medium * 0.025
        leads_optimistic = total_medium * 0.04
        
        print(f"🟡 MEDIUM PRIORITY PAGES (Charges + Notaire):")
        print(f"   Combined monthly clicks: {total_medium:.0f}")
        print(f"   Estimated leads range: {leads_conservative:.0f} - {leads_optimistic:.0f}/month")
        print(f"   Realistic estimate: {leads_realistic:.0f} leads/month")
        print(f"   At 45€/lead: €{leads_realistic * 45:,.0f}/month revenue potential\n")
    
    # Get keywords related to credit/financing
    print("="*180)
    print("🔍 SEARCH CONSOLE KEYWORDS - Credit/Financing Intent")
    print("="*180 + "\n")
    
    if 'Requêtes' in wb.sheetnames:
        ws_req = wb['Requêtes']
        
        credit_keywords = []
        
        for row in ws_req.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                break
            
            query = str(row[0]).lower()
            clics = row[1] or 0
            impr = row[2] or 0
            
            # Match credit-related keywords
            if any(term in query for term in ['crédit', 'prêt', 'financement', 'emprunt', 'taux', 'rachat', 'consolidation', 'travaux']):
                credit_keywords.append({
                    'query': row[0],
                    'clics': clics,
                    'impr': impr,
                })
        
        # Sort by clicks
        credit_keywords.sort(key=lambda x: x['clics'], reverse=True)
        
        if credit_keywords:
            print(f"Found {len(credit_keywords)} credit-related keywords:\n")
            for i, kw in enumerate(credit_keywords[:15], 1):
                print(f"{i:2d}. \"{kw['query']}\" | {kw['clics']:.0f} clics | {kw['impr']:.0f} impr")
        else:
            print("No credit-related keywords found in current data")
    
    print("\n" + "="*180)
    print("✅ Analysis complete!")
    print("="*180 + "\n")
    
except Exception as e:
    import traceback
    print(f"❌ Error reading Excel file: {e}")
    traceback.print_exc()
